"""Tests for scripts/import_experts.py — validation logic."""
import csv
import sys
from pathlib import Path

import openpyxl
import pytest

# Make scripts importable
_SCRIPTS = Path(__file__).resolve().parent.parent / "scripts"
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

from import_experts import run  # noqa: E402


# ── helpers ───────────────────────────────────────────────────────────────────

_VALID_GROUP_TEAMS = [
    # Group A
    ("A", "Mexico", "South Africa"),
    ("A", "Mexico", "Korea Republic"),
    ("A", "Mexico", "Czechia"),
    ("A", "South Africa", "Korea Republic"),
    ("A", "South Africa", "Czechia"),
    ("A", "Korea Republic", "Czechia"),
]


def _make_valid_xlsx(tmp_path: Path, rows=None, row_count: int = 72) -> Path:
    """Write a minimal valid xlsx with canonical English team names."""
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Group", "Team_A", "Team_B", "Davo", None, "Maldini", None, "Promedio", None])

    # Use the actual 72-row fixture from the real Excel if available
    real_xlsx = Path.home() / "Downloads" / "Poia.xlsx"
    if rows is None and real_xlsx.exists():
        real_wb = openpyxl.load_workbook(str(real_xlsx), data_only=True)
        real_ws = real_wb["Sheet1"]
        real_rows = [r for r in real_ws.iter_rows(values_only=True)][1:]
        rows = [r[:7] for r in real_rows[:row_count]]
    elif rows is None:
        # Fallback: generate synthetic valid rows
        import json
        from pathlib import Path as P
        fixtures_path = P(__file__).resolve().parent.parent / "data" / "fixtures.json"
        with open(fixtures_path, encoding="utf-8") as f:
            all_fx = json.load(f)
        group_fx = [fx for fx in all_fx if fx["stage"].startswith("Group")]
        rows = []
        for fx in group_fx[:row_count]:
            letter = fx["stage"].split()[-1]
            rows.append((letter, fx["home"], fx["away"], 1, 0, 2, 0))

    for r in rows:
        ws.append(list(r))
    wb.save(p)
    return p


# ── valid round-trip ──────────────────────────────────────────────────────────

def test_valid_file_writes_csv(tmp_path):
    """Valid xlsx -> experts.csv with 72 rows and correct columns."""
    xlsx = _make_valid_xlsx(tmp_path)
    out = tmp_path / "experts.csv"
    import import_experts
    original_out = import_experts.OUT_CSV
    import_experts.OUT_CSV = out
    try:
        run(xlsx)
    finally:
        import_experts.OUT_CSV = original_out

    assert out.exists()
    with open(out, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 72
    expected_cols = {"group", "team_a", "team_b",
                     "davo_a", "davo_b", "maldini_a", "maldini_b"}
    assert set(rows[0].keys()) == expected_cols


# ── wrong row count ───────────────────────────────────────────────────────────

def test_wrong_row_count_fails(tmp_path):
    """xlsx with != 72 data rows should sys.exit(1)."""
    xlsx = _make_valid_xlsx(tmp_path, row_count=10)
    out = tmp_path / "experts.csv"
    import import_experts
    original_out = import_experts.OUT_CSV
    import_experts.OUT_CSV = out
    try:
        with pytest.raises(SystemExit) as exc_info:
            run(xlsx)
        assert exc_info.value.code == 1
        assert not out.exists()  # never written partial output
    finally:
        import_experts.OUT_CSV = original_out


# ── bad team name ─────────────────────────────────────────────────────────────

def test_bad_team_name_fails(tmp_path):
    """A team name that doesn't resolve should exit(1) with the name listed."""
    real_xlsx = Path.home() / "Downloads" / "Poia.xlsx"
    if not real_xlsx.exists():
        pytest.skip("Poia.xlsx not available")

    real_wb = openpyxl.load_workbook(str(real_xlsx), data_only=True)
    real_ws = real_wb["Sheet1"]
    real_rows = [list(r[:7]) for r in real_ws.iter_rows(values_only=True)][1:]

    # Corrupt one team name in the first row
    real_rows[0][1] = "NotARealTeam"

    xlsx = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Group", "Team_A", "Team_B", "Davo", None, "Maldini", None])
    for r in real_rows:
        ws.append(r)
    wb.save(xlsx)

    out = tmp_path / "experts.csv"
    import import_experts
    original_out = import_experts.OUT_CSV
    import_experts.OUT_CSV = out
    try:
        with pytest.raises(SystemExit) as exc_info:
            run(xlsx)
        assert exc_info.value.code == 1
        assert not out.exists()
    finally:
        import_experts.OUT_CSV = original_out


# ── full coverage: all 72 rows match all 72 group fixtures ───────────────────

def test_all_72_rows_match_group_fixtures():
    """All 72 expert rows must match exactly one group-stage fixture (either orientation).

    This is the definitive coverage test for the alias table.
    """
    import json
    from data.experts import load_experts, picks_for
    from pathlib import Path as P

    fixtures_path = P(__file__).resolve().parent.parent / "data" / "fixtures.json"
    with open(fixtures_path, encoding="utf-8") as f:
        all_fx = json.load(f)
    group_fixtures = [fx for fx in all_fx if fx["stage"].startswith("Group")]

    csv_path = P(__file__).resolve().parent.parent / "data" / "experts.csv"
    if not csv_path.exists():
        pytest.skip("data/experts.csv not present")

    from unittest.mock import patch
    with patch("data.experts._CSV_PATH", csv_path):
        experts = load_experts()

    matched = 0
    swapped = 0
    unmatched = []

    for fx in group_fixtures:
        home, away = fx["home"], fx["away"]
        result = picks_for(experts, home, away)
        if result is None:
            unmatched.append(f"{home} vs {away}")
        else:
            matched += 1
            # Check if it was swapped by looking up directly
            from data.team_names import same_team
            found_direct = False
            for (ta, tb) in experts.keys():
                if same_team(home, ta) and same_team(away, tb):
                    found_direct = True
                    break
            if not found_direct:
                swapped += 1

    if unmatched:
        pytest.fail(
            f"{len(unmatched)} fixtures had no expert match:\n"
            + "\n".join(f"  {m}" for m in unmatched)
        )

    assert matched == 72, f"Expected 72 matches, got {matched}"
    # Report swapped count (not an error condition, just informational)
    print(f"\nFull coverage: {matched}/72 matched, {swapped} orientation-swapped")
