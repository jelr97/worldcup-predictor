"""Import expert score predictions from Poia.xlsx into data/experts.csv.

Usage:
    python scripts/import_experts.py [path/to/Poia.xlsx]

Default path: ~/Downloads/Poia.xlsx

Validates:
- Exactly 72 data rows
- Integer scores 0-15
- Every team name resolves (via team_names.normalize) to a group-stage fixture team
Fails loudly listing every unmatched name; never writes partial output.
"""
import csv
import sys
from pathlib import Path

# Allow running from repo root or scripts/ dir
_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import openpyxl  # noqa: E402

from data.fixtures import load_fixtures  # noqa: E402
from data.team_names import normalize, same_team  # noqa: E402

EXPECTED_ROWS = 72
SCORE_MIN, SCORE_MAX = 0, 15
DEFAULT_XLSX = Path.home() / "Downloads" / "Poia.xlsx"
OUT_CSV = _ROOT / "data" / "experts.csv"


def _group_stage_teams(fixtures):
    """Return the set of fixture team names (original, not normalized) for group stage."""
    teams = []
    for fx in fixtures:
        if fx["stage"].startswith("Group"):
            teams.append(fx["home"])
            teams.append(fx["away"])
    return teams


def _resolve_team(name: str, fixture_teams: list[str]) -> str | None:
    """Return the matching fixture team name if found, else None."""
    for ft in fixture_teams:
        if same_team(name, ft):
            return ft
    return None


def run(xlsx_path: Path = DEFAULT_XLSX) -> None:
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if "Sheet1" not in wb.sheetnames:
        print("ERROR: sheet 'Sheet1' not found in workbook", file=sys.stderr)
        sys.exit(1)

    ws = wb["Sheet1"]
    rows = [r for r in ws.iter_rows(values_only=True)][1:]  # skip header

    # ── row count ────────────────────────────────────────────────────────────────
    if len(rows) != EXPECTED_ROWS:
        print(
            f"ERROR: expected {EXPECTED_ROWS} data rows, got {len(rows)}",
            file=sys.stderr,
        )
        sys.exit(1)

    # ── fixture team lookup set ───────────────────────────────────────────────
    fixtures = load_fixtures()
    fixture_teams = _group_stage_teams(fixtures)

    # ── validate every row ───────────────────────────────────────────────────
    errors: list[str] = []
    parsed: list[dict] = []

    for i, row in enumerate(rows, start=2):  # row 2 is first data row in Excel
        group = row[0]
        team_a_raw = str(row[1]) if row[1] is not None else ""
        team_b_raw = str(row[2]) if row[2] is not None else ""
        davo_a = row[3]
        davo_b = row[4]
        maldini_a = row[5]
        maldini_b = row[6]

        # Resolve teams
        ft_a = _resolve_team(team_a_raw, fixture_teams)
        ft_b = _resolve_team(team_b_raw, fixture_teams)
        if ft_a is None:
            errors.append(f"Row {i}: unmatched team name {team_a_raw!r}")
        if ft_b is None:
            errors.append(f"Row {i}: unmatched team name {team_b_raw!r}")

        # Validate scores are integers in [0, 15]
        for label, val in [
            (f"Row {i} davo_a", davo_a),
            (f"Row {i} davo_b", davo_b),
            (f"Row {i} maldini_a", maldini_a),
            (f"Row {i} maldini_b", maldini_b),
        ]:
            if not isinstance(val, (int, float)) or int(val) != val:
                errors.append(f"{label}: expected integer, got {val!r}")
            elif not (SCORE_MIN <= int(val) <= SCORE_MAX):
                errors.append(
                    f"{label}: score {val} out of range [{SCORE_MIN}, {SCORE_MAX}]"
                )

        if ft_a and ft_b and not errors:
            parsed.append({
                "group": str(group),
                "team_a": ft_a,
                "team_b": ft_b,
                "davo_a": int(davo_a),
                "davo_b": int(davo_b),
                "maldini_a": int(maldini_a),
                "maldini_b": int(maldini_b),
            })

    if errors:
        print("VALIDATION ERRORS — not writing output:", file=sys.stderr)
        for e in errors:
            print(f"  {e}", file=sys.stderr)
        sys.exit(1)

    # ── write CSV ─────────────────────────────────────────────────────────────
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["group", "team_a", "team_b",
                        "davo_a", "davo_b", "maldini_a", "maldini_b"],
        )
        writer.writeheader()
        writer.writerows(parsed)

    print(f"Wrote {len(parsed)} rows to {OUT_CSV}")


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    run(path)
